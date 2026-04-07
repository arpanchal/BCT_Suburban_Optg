import streamlit as st
import os, sys, json, re, shutil
import datetime as _dt_mod
from datetime import datetime
from collections import defaultdict

# Aliases for isinstance checks on openpyxl cell values
_dt_time     = _dt_mod.time
_dt_datetime = _dt_mod.datetime

st.set_page_config(page_title="Regenerate Data", layout="wide")
st.title("🔄 Regenerate Data from Excel")
st.caption("Upload an updated WTT Excel file to rebuild all JSON data files used by the app.")

BASE = os.path.dirname(os.path.dirname(__file__))

CORRECT_SEQ = [
    'CCG','MEL','CYR','GTR','BCL','BCT','MX','PL','PBHD','DDR',
    'MRU','MM','BA','BDTS','KHAR','STC','VLP','ADH',
    'JOS','RMAR','GMN','MDD','KILE','BVI',
    'DIC','MIRA','BYR','NIG','BSR','NSP','VR',
    'VTN','SAH','KLV','PLG','UOI','BOR','VGN','DRD'
]
ARR_DEP_BASES = {
    'DDR','ADH','BVI','BYR','BSR','NSP','VR',
    'VTN','SAH','KLV','PLG','UOI','BOR','VGN','DRD'
}
ALL_RAW_STNS  = set(CORRECT_SEQ) | {f'{s} ARR' for s in ARR_DEP_BASES} | {f'{s} DEP' for s in ARR_DEP_BASES}

# ── Show current data status ──────────────────────────────────────────────────
st.markdown("### Current Data Status")
stops_path = os.path.join(BASE, "stops_data.json")
meta_path  = os.path.join(BASE, "train_meta.json")

col1, col2, col3 = st.columns(3)
with col1:
    if os.path.exists(stops_path):
        mtime = datetime.fromtimestamp(os.path.getmtime(stops_path))
        with open(stops_path) as f:
            sd = json.load(f)
        st.success(f"✅ stops_data.json\n\n"
                   f"**{len(sd['stops']):,} stops** · {len(sd['stations'])} stations\n\n"
                   f"Last updated: {mtime.strftime('%d %b %Y %H:%M')}")
    else:
        st.error("❌ stops_data.json — not found")

with col2:
    if os.path.exists(meta_path):
        mtime = datetime.fromtimestamp(os.path.getmtime(meta_path))
        with open(meta_path) as f:
            md = json.load(f)
        st.success(f"✅ train_meta.json\n\n"
                   f"**{len(md):,} trains**\n\n"
                   f"Last updated: {mtime.strftime('%d %b %Y %H:%M')}")
    else:
        st.error("❌ train_meta.json — not found")

with col3:
    st.info("📋 **Station sequence**\n\n" + " → ".join(CORRECT_SEQ[:10]) + "\n→ …(" + str(len(CORRECT_SEQ)) + " total)")

st.markdown("---")
st.markdown("### Upload Updated Excel File")
st.markdown(
    "Upload the WTT Excel file (same format as `wtt_mar26_formatted.xlsx`). "
    "The app expects:\n"
    "- **Sheet 1** — Train Summary (Train No., Set No., Direction, Type, AC, DRD, From, To, Link, Cars, Platform, …)\n"
    "- **Sheet 2** — Stop Schedule (reference matrix, not used for data extraction)\n""- **Sheet 7** — Stop Details (flat list: Train No., Direction, Type, AC, From, To, Station, Time, **Line**)\n"
)

uploaded = st.file_uploader("Upload Excel (.xlsx)", type=["xlsx"])

if uploaded:
    st.info(f"File received: **{uploaded.name}** ({uploaded.size/1024:.1f} KB)")

    if st.button("🚀 Regenerate JSON Data Files", type="primary"):
        progress = st.progress(0, text="Starting…")
        log      = st.empty()
        logs     = []

        def logit(msg):
            logs.append(msg)
            log.code("\n".join(logs))

        try:
            from openpyxl import load_workbook
            import tempfile

            # Save uploaded file to temp location
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(uploaded.read())
                tmp_path = tmp.name
            logit(f"✓ Saved uploaded file to temp: {tmp_path}")
            progress.progress(5, "Loading workbook…")

            wb     = load_workbook(tmp_path, data_only=True)
            ws_sum = wb['Train Summary']
            ws_det = wb['Stop Details']
            logit(f"✓ Workbook loaded. Sheets: {wb.sheetnames}")
            progress.progress(15, "Reading train metadata from Train Summary…")

            # ── Train Summary metadata ────────────────────────────────────────
            train_info = {}
            for row in ws_sum.iter_rows(min_row=3, values_only=True):
                tr = row[0]
                if tr is None: break
                tr_s = str(tr).strip()

                def _pt(v):
                    if isinstance(v, (_dt_time, _dt_datetime)):
                        return f"{v.hour:02d}:{v.minute:02d}"
                    if v:
                        m2 = re.match(r'^(\d{1,2}):(\d{2})(?::\d{2})?$', str(v).strip())
                        if m2: return f"{int(m2.group(1)):02d}:{int(m2.group(2)):02d}"
                    return ''

                typ_s = str(row[3] or '').strip()
                ac_s  = str(row[4] or '').strip()
                drd_s = str(row[5] or '').strip()
                dir_s = str(row[2] or '').strip()

                # Normalize direction variants
                if dir_s.upper() in ('DN', 'D', 'DOWN'):  dir_s = 'DOWN'
                elif dir_s.upper() in ('UP', 'U'):        dir_s = 'UP'

                if tr_s.startswith('ETY'):                      _cat = 'EMPTY'
                elif typ_s == 'M/E':                            _cat = 'MAIL'
                elif drd_s == 'DRD' or tr_s.startswith('93'):  _cat = 'DRD'
                elif ac_s == 'AC' or tr_s.startswith('94'):    _cat = 'AC_LOCAL'
                else:                                            _cat = 'LOCAL'

                train_info[tr_s] = {
                    'set_no': str(row[1] or ''), 'direction': dir_s,
                    'type': typ_s, 'ac': ac_s, 'drd': drd_s,
                    'from_stn': str(row[6] or ''), 'to_stn': str(row[7] or ''),
                    'link': str(row[8] or ''), 'cars': str(row[9] or ''),
                    'platform': str(row[10] or ''),
                    'first_dep': _pt(row[11]), 'last_arr': _pt(row[12]),
                    'dep_time':  _pt(row[13]),
                    'rev_as': str(row[14] or ''), 'rev_set_no': str(row[15] or ''),
                    'train_cat': _cat, 'line': '', 'line_switch': {},
                }
            logit(f"✓ Train metadata: {len(train_info)} trains")
            progress.progress(30, "Parsing Stop Details sheet…")

            # ── Stop Details parsing ──────────────────────────────────────────
            # Columns: Train No., Direction, Type, AC, From, To, Station, Time, Line
            KNOWN_STNS = set(CORRECT_SEQ) | {'PNVL'}
            for _b in ARR_DEP_BASES:
                KNOWN_STNS.add(f'{_b} ARR')
                KNOWN_STNS.add(f'{_b} DEP')

            ARR_VARS = {f'{b} ARR' for b in ARR_DEP_BASES}
            DEP_VARS = {f'{b} DEP' for b in ARR_DEP_BASES}

            raw_by_train = defaultdict(list)
            skipped = 0

            for row in ws_det.iter_rows(min_row=3, values_only=True):
                tr, dir_, typ, ac, frm, to, stn, t, line = row
                if tr is None: continue
                tr_s = str(tr).strip()
                if ',' in tr_s: skipped += 1; continue          # multi-train row
                # Allow numeric, numeric+A suffix (e.g. 90007A), ETY names
                if not re.match(r'^(\d+A?|ETY.*)$', tr_s): skipped += 1; continue
                # Normalize direction
                if dir_:
                    d_s = str(dir_).strip().upper()
                    if d_s in ('DN', 'D', 'DOWN'): dir_ = 'DOWN'
                    elif d_s in ('UP', 'U'):        dir_ = 'UP'
                stn_s = str(stn).strip() if stn else ''
                if stn_s not in KNOWN_STNS: skipped += 1; continue  # unknown station

                # Parse time
                if isinstance(t, (_dt_time, _dt_datetime)):
                    h, mn = t.hour, t.minute
                elif isinstance(t, str):
                    m2 = re.match(r'^(\d{1,2}):(\d{2})(?::\d{2})?$', t.strip())
                    if not m2: skipped += 1; continue
                    h, mn = int(m2.group(1)), int(m2.group(2))
                else:
                    skipped += 1; continue

                info = train_info.get(tr_s, {})
                raw_by_train[tr_s].append({
                    'stn': stn_s,
                    'time': f"{h:02d}:{mn:02d}",
                    'minutes': h*60 + mn,
                    'direction': str(dir_ or info.get('direction','')).strip(),
                    'type':      str(typ  or info.get('type','')).strip(),
                    'ac':        str(ac   or info.get('ac','')  ).strip(),
                    'from_stn':  str(frm  or info.get('from_stn','')).strip(),
                    'to_stn':    str(to   or info.get('to_stn', '')).strip(),
                    'line':      str(line or '').strip(),
                    'train_cat': info.get('train_cat','LOCAL'),
                })

            logit(f"✓ Stop Details parsed: {sum(len(v) for v in raw_by_train.values())} records "
                  f"from {len(raw_by_train)} trains ({skipped} rows skipped)")
            progress.progress(60, "Fixing overnight trains & building stops…")

            # ── Overnight fix ─────────────────────────────────────────────────
            def fix_overnight(pts):
                has_late  = any(p['minutes'] >= 22*60 for p in pts)
                has_early = any(p['minutes'] <   2*60 for p in pts)
                if has_late and has_early:
                    return [{**p, 'minutes': p['minutes']+1440}
                            if p['minutes'] < 2*60 else p for p in pts]
                return pts

            # ── Build final stops ─────────────────────────────────────────────
            final_stops = []
            for tr_s, pts in raw_by_train.items():
                pts = fix_overnight(pts)

                halt_pts = []
                for p in pts:
                    stn = p['stn']
                    if stn in ARR_VARS:
                        halt_pts.append({**p, '_role':'arr', '_base':stn.replace(' ARR','')})
                    elif stn in DEP_VARS:
                        halt_pts.append({**p, '_role':'dep', '_base':stn.replace(' DEP','')})
                    elif stn in ARR_DEP_BASES:
                        halt_pts.append({**p, '_role':'pass', '_base':stn})
                    else:
                        final_stops.append({
                            'train': tr_s, 'station': stn, 'time': p['time'],
                            'minutes': p['minutes'], 'direction': p['direction'],
                            'type': p['type'], 'ac': p['ac'],
                            'from_stn': p['from_stn'], 'to_stn': p['to_stn'],
                            'stop_type': 'stop', 'train_cat': p['train_cat'],
                            'line': p['line'],
                        })

                # Group ARR/DEP by base station
                by_base = defaultdict(dict)
                for p in halt_pts:
                    by_base[p['_base']][p['_role']] = p

                for base, roles in by_base.items():
                    ref = roles.get('arr') or roles.get('dep') or roles.get('pass')
                    rec = {
                        'train': tr_s, 'station': base,
                        'direction': ref['direction'], 'type': ref['type'],
                        'ac': ref['ac'], 'from_stn': ref['from_stn'],
                        'to_stn': ref['to_stn'], 'train_cat': ref['train_cat'],
                        'line': ref['line'],
                    }
                    if 'arr' in roles and 'dep' in roles:
                        a, d = roles['arr'], roles['dep']
                        rec.update({'stop_type':'dwell',
                                    'arr_time':a['time'],'arr_minutes':a['minutes'],
                                    'dep_time':d['time'],'dep_minutes':d['minutes'],
                                    'time':a['time'],'minutes':a['minutes']})
                    elif 'dep' in roles:
                        d = roles['dep']
                        rec.update({'stop_type':'dep','dep_time':d['time'],
                                    'dep_minutes':d['minutes'],
                                    'time':d['time'],'minutes':d['minutes']})
                    elif 'arr' in roles:
                        a = roles['arr']
                        rec.update({'stop_type':'arr','arr_time':a['time'],
                                    'arr_minutes':a['minutes'],
                                    'time':a['time'],'minutes':a['minutes']})
                    else:
                        p = roles['pass']
                        rec.update({'stop_type':'pass',
                                    'time':p['time'],'minutes':p['minutes']})
                    final_stops.append(rec)

            

            logit(f"✓ Total stops built: {len(final_stops)}")
            progress.progress(80, "Saving JSON files…")

            # ── Backup existing files ─────────────────────────────────────────
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            for fname in ["stops_data.json","train_meta.json"]:
                fpath = os.path.join(BASE, fname)
                if os.path.exists(fpath):
                    backup = os.path.join(BASE, f"{fname}.backup_{ts}")
                    shutil.copy2(fpath, backup)
                    logit(f"✓ Backed up {fname} → {os.path.basename(backup)}")

            # ── Write new JSON files ──────────────────────────────────────────
            with open(stops_path, 'w') as f:
                json.dump({'stops': final_stops, 'stations': CORRECT_SEQ,
                           'arr_dep_bases': list(ARR_DEP_BASES)}, f)
            logit(f"✓ stops_data.json written ({os.path.getsize(stops_path)/1024:.0f} KB)")

            with open(meta_path, 'w') as f:
                json.dump(train_info, f)
            logit(f"✓ train_meta.json written ({os.path.getsize(meta_path)/1024:.0f} KB)")

            os.unlink(tmp_path)

            # ── Clear Streamlit cache so all pages reload fresh data ──────────
            st.cache_data.clear()
            progress.progress(100, "✅ Done!")

            # ── Validation summary ────────────────────────────────────────────
            from collections import Counter
            st.success("✅ Data regenerated successfully!")
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("Total stops",   f"{len(final_stops):,}")
            c2.metric("Trains",        f"{len(train_info):,}")
            c3.metric("Stations",      str(len(CORRECT_SEQ)))
            stypes = Counter(s.get('stop_type','stop') for s in final_stops)
            c4.metric("Dwell records", str(stypes.get('dwell',0)))

            overnight_count = sum(1 for s in final_stops if s['minutes'] >= 1440)
            st.info(f"🌙 Overnight stops (past midnight): **{overnight_count}**")
            st.success("✅ All pages will now automatically load the new data — no app restart needed!")

        except Exception as e:
            progress.progress(0, "Error")
            st.error(f"❌ Error during regeneration: {e}")
            logit(f"ERROR: {e}")
            import traceback
            st.code(traceback.format_exc())

# ── Manual regeneration from bundled Excel ────────────────────────────────────
st.markdown("---")
st.markdown("### Re-generate from Bundled Excel")
st.caption("If the stops_data.json is missing or corrupted, regenerate from the original Excel shipped with the app.")

bundled_excel = os.path.join(BASE, "wtt_mar26_updated.xlsx")
if os.path.exists(bundled_excel):
    st.info(f"Found bundled Excel: `wtt_mar26_updated.xlsx` ({os.path.getsize(bundled_excel)/1024:.0f} KB)")
    if st.button("Regenerate from bundled Excel"):
        with open(bundled_excel, "rb") as f:
            data = f.read()
        import io
        class FakeUpload:
            def __init__(self, data, name):
                self.name = name
                self.size = len(data)
                self._data = data
            def read(self): return self._data
        st.warning("Use the uploader above — paste wtt_mar26_updated.xlsx as your upload.")
else:
    st.info("Place `wtt_mar26_updated.xlsx` in the app folder to enable bundled regeneration.")

# ── Station sequence editor ───────────────────────────────────────────────────
st.markdown("---")
with st.expander("⚙️ Advanced: Edit Station Sequence"):
    st.caption("Edit the station sequence used for Y-axis ordering. Changes take effect on next data regeneration.")
    current_seq = " → ".join(CORRECT_SEQ)
    st.markdown(f"**Current sequence ({len(CORRECT_SEQ)} stations):**")
    st.code(current_seq)
    st.info("To change the station sequence, edit `CORRECT_SEQ` in `pages/0_Regenerate_Data.py` "
            "and `pages/utils.py`, then regenerate data.")
